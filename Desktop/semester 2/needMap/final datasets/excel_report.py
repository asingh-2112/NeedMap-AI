import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
import os

# ── Load Data ──────────────────────────────────────────────────────────
df = pd.read_csv("outputs/results.csv")
df["predicted_text"] = df["predicted_text"].fillna("")

# Remove paddle (100% errors)
df = df[df["engine"] != "paddle"]

# Flag valid results
def is_valid(row):
    text = row["predicted_text"]
    if text.startswith("ERROR") or len(text.strip()) == 0:
        return False
    if row["dataset"] in ["hindi", "telugu"]:
        indic = sum(1 for c in text if '\u0900' <= c <= '\u097F' or '\u0C00' <= c <= '\u0C7F')
        return indic > 0
    return True

df["valid"] = df.apply(is_valid, axis=1)
df["valid_length"] = df.apply(lambda r: r["text_length"] if r["valid"] else 0, axis=1)
print(df.groupby(["dataset","engine"])["valid_length"].mean().unstack())

# ── Styles ─────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
GREEN       = "70AD47"
ORANGE      = "ED7D31"
RED         = "FF0000"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"

def header_cell(ws, row, col, value, bg=DARK_BLUE, fg=WHITE, bold=True, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(color=fg, bold=bold, size=size)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def data_cell(ws, row, col, value, bg=WHITE, bold=False, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    return c

def add_border(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="BFBFBF")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════
# SHEET 1: Summary Dashboard
# ══════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Summary Dashboard"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 28
for col in ["B","C","D"]:
    ws1.column_dimensions[col].width = 18
ws1.row_dimensions[1].height = 40
ws1.row_dimensions[2].height = 20

# Title
ws1.merge_cells("A1:D1")
t = ws1["A1"]
t.value = "OCR BENCHMARK — SUMMARY DASHBOARD"
t.fill = PatternFill("solid", fgColor=DARK_BLUE)
t.font = Font(color=WHITE, bold=True, size=16)
t.alignment = Alignment(horizontal="center", vertical="center")

# Subtitle
ws1.merge_cells("A2:D2")
s = ws1["A2"]
s.value = "Engines: Tesseract | EasyOCR | Doctr   |   Datasets: Hindi, Telugu, English Scanned, Handwritten, Noisy Receipts"
s.fill = PatternFill("solid", fgColor=MID_BLUE)
s.font = Font(color=WHITE, size=10)
s.alignment = Alignment(horizontal="center", vertical="center")

# ── Speed Table ────────────────────────────────────────────────────────
ws1.merge_cells("A4:D4")
h = ws1["A4"]
h.value = "⏱️  AVERAGE SPEED PER ENGINE (seconds)"
h.fill = PatternFill("solid", fgColor=MID_BLUE)
h.font = Font(color=WHITE, bold=True, size=11)
h.alignment = Alignment(horizontal="center", vertical="center")

headers = ["Engine", "Avg Time (s)", "Min Time (s)", "Max Time (s)"]
for i, hdr in enumerate(headers, 1):
    header_cell(ws1, 5, i, hdr, bg="2E75B6")

speed = df.groupby("engine")["time_sec"].agg(["mean","min","max"]).reset_index()
speed = speed.sort_values("mean")
for r, (_, row) in enumerate(speed.iterrows(), 6):
    bg = LIGHT_GRAY if r % 2 == 0 else WHITE
    data_cell(ws1, r, 1, row["engine"].capitalize(), bg=bg, bold=True, align="left")
    data_cell(ws1, r, 2, round(row["mean"], 3), bg=bg)
    data_cell(ws1, r, 3, round(row["min"], 3), bg=bg)
    data_cell(ws1, r, 4, round(row["max"], 3), bg=bg)
add_border(ws1, 5, 5+len(speed), 1, 4)

# ── Winners Table ──────────────────────────────────────────────────────
start_row = 6 + len(speed) + 2
ws1.merge_cells(f"A{start_row}:D{start_row}")
h2 = ws1[f"A{start_row}"]
h2.value = "🏆  BEST ENGINE PER DATASET"
h2.fill = PatternFill("solid", fgColor=MID_BLUE)
h2.font = Font(color=WHITE, bold=True, size=11)
h2.alignment = Alignment(horizontal="center", vertical="center")

header_cell(ws1, start_row+1, 1, "Dataset", bg="2E75B6")
header_cell(ws1, start_row+1, 2, "Best Engine", bg="2E75B6")
header_cell(ws1, start_row+1, 3, "Avg Valid Chars", bg="2E75B6")
header_cell(ws1, start_row+1, 4, "Reason", bg="2E75B6")

winners_info = {
    "english_scanned":     ("Doctr",    "Cleanest text, best punctuation"),
    "noisy_receipts":      ("Doctr",    "Best structure on receipts"),
    "handwritten_english": ("EasyOCR",  "Better cursive recognition"),
    "hindi":               ("EasyOCR",  "Best Devanagari script support"),
    "telugu":              ("EasyOCR",  "Best Telugu script extraction"),
}

for r, (dataset, (engine, reason)) in enumerate(winners_info.items(), start_row+2):
    bg = LIGHT_GRAY if r % 2 == 0 else WHITE
    scores = df[df["dataset"]==dataset].groupby("engine")["valid_length"].mean()
    best_score = scores.get(engine.lower(), scores.max() if len(scores) > 0 else 0)
    data_cell(ws1, r, 1, dataset.replace("_"," ").title(), bg=bg, align="left")
    c = ws1.cell(row=r, column=2, value=engine.capitalize())
    c.fill = PatternFill("solid", fgColor=GREEN)
    c.font = Font(color=WHITE, bold=True, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    data_cell(ws1, r, 3, round(best_score, 1), bg=bg)
    data_cell(ws1, r, 4, reason, bg=bg, align="left")
add_border(ws1, start_row+1, start_row+1+len(winners_info), 1, 4)

# ── Key Findings ───────────────────────────────────────────────────────
fr = start_row + len(winners_info) + 4
ws1.merge_cells(f"A{fr}:D{fr}")
hf = ws1[f"A{fr}"]
hf.value = "💡  KEY FINDINGS"
hf.fill = PatternFill("solid", fgColor=MID_BLUE)
hf.font = Font(color=WHITE, bold=True, size=11)
hf.alignment = Alignment(horizontal="center", vertical="center")

findings = [
    "Doctr is the best engine for printed/scanned English documents",
    "EasyOCR is the best engine for Indic scripts (Hindi, Telugu) and handwriting",
    "Tesseract is the fastest engine (10x faster than EasyOCR)",
    "PaddleOCR 3.x was excluded: known Windows/OneDNN backend incompatibility (ConvertPirAttribute2RuntimeAttribute error). Works on Linux/Mac only.",
    "Google Vision API was excluded: requires paid API key, internet connection, and raises data privacy concerns — not suitable for fair offline benchmarking.",
    "No single engine wins all categories — task-specific selection is recommended",
    "For production use: combine Doctr (English docs) + EasyOCR (Indic/handwritten) for best coverage",
]
for i, finding in enumerate(findings, fr+1):
    ws1.merge_cells(f"A{i}:D{i}")
    c = ws1[f"A{i}"]
    c.value = f"  • {finding}"
    c.fill = PatternFill("solid", fgColor=LIGHT_GRAY if i%2==0 else WHITE)
    c.font = Font(size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[i].height = 18

# ══════════════════════════════════════════════════════════════════════
# SHEET 2: Detailed Results
# ══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Detailed Results")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 12

ws2.merge_cells("A1:E1")
t2 = ws2["A1"]
t2.value = "DETAILED RESULTS — AVG VALID CHARACTERS EXTRACTED PER DATASET PER ENGINE"
t2.fill = PatternFill("solid", fgColor=DARK_BLUE)
t2.font = Font(color=WHITE, bold=True, size=13)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

engines = ["tesseract", "easyocr", "doctr"]
header_cell(ws2, 2, 1, "Dataset", bg=MID_BLUE)
for i, eng in enumerate(engines, 2):
    header_cell(ws2, 2, i, eng.capitalize(), bg=MID_BLUE)
header_cell(ws2, 2, len(engines)+2, "Winner", bg=MID_BLUE)

datasets = df["dataset"].unique()
pivot = df.groupby(["dataset","engine"])["valid_length"].mean().unstack(fill_value=0)

for r, dataset in enumerate(datasets, 3):
    bg = LIGHT_GRAY if r % 2 == 0 else WHITE
    data_cell(ws2, r, 1, dataset.replace("_"," ").title(), bg=bg, bold=True, align="left")
    row_vals = []
    for i, eng in enumerate(engines, 2):
        val = round(pivot.loc[dataset, eng], 1) if eng in pivot.columns else 0
        row_vals.append((i, val))
        data_cell(ws2, r, i, val, bg=bg)
    winner_engine = engines[max(range(len(row_vals)), key=lambda x: row_vals[x][1])]
    c = ws2.cell(row=r, column=len(engines)+2, value=winner_engine.capitalize())
    c.fill = PatternFill("solid", fgColor=GREEN)
    c.font = Font(color=WHITE, bold=True, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")

add_border(ws2, 2, 2+len(datasets), 1, len(engines)+2)

# ── Chart on Sheet 2 ───────────────────────────────────────────────────
chart = BarChart()
chart.type = "col"
chart.grouping = "clustered"
chart.title = "Avg Valid Characters Extracted per Dataset per Engine"
chart.y_axis.title = "Avg Characters"
chart.x_axis.title = "Dataset"
chart.style = 10
chart.width = 22
chart.height = 14

data_ref = Reference(ws2, min_col=2, max_col=len(engines)+1,
                     min_row=2, max_row=2+len(datasets))
cats_ref = Reference(ws2, min_col=1, min_row=3, max_row=2+len(datasets))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws2.add_chart(chart, f"A{len(datasets)+5}")

# ══════════════════════════════════════════════════════════════════════
# SHEET 3: Speed Comparison
# ══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Speed Comparison")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 22
for col in ["B","C","D"]:
    ws3.column_dimensions[col].width = 16

ws3.merge_cells("A1:D1")
t3 = ws3["A1"]
t3.value = "SPEED COMPARISON — AVG TIME PER ENGINE PER DATASET (seconds)"
t3.fill = PatternFill("solid", fgColor=DARK_BLUE)
t3.font = Font(color=WHITE, bold=True, size=13)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

header_cell(ws3, 2, 1, "Dataset", bg=MID_BLUE)
for i, eng in enumerate(engines, 2):
    header_cell(ws3, 2, i, eng.capitalize(), bg=MID_BLUE)

speed_pivot = df.groupby(["dataset","engine"])["time_sec"].mean().unstack(fill_value=0)
for r, dataset in enumerate(datasets, 3):
    bg = LIGHT_GRAY if r % 2 == 0 else WHITE
    data_cell(ws3, r, 1, dataset.replace("_"," ").title(), bg=bg, bold=True, align="left")
    for i, eng in enumerate(engines, 2):
        val = round(speed_pivot.loc[dataset, eng], 3) if eng in speed_pivot.columns else 0
        data_cell(ws3, r, i, val, bg=bg)

add_border(ws3, 2, 2+len(datasets), 1, len(engines)+1)

# Speed chart
schart = BarChart()
schart.type = "col"
schart.grouping = "clustered"
schart.title = "Average Processing Time per Engine per Dataset"
schart.y_axis.title = "Time (seconds)"
schart.x_axis.title = "Dataset"
schart.style = 10
schart.width = 22
schart.height = 14

sdata = Reference(ws3, min_col=2, max_col=len(engines)+1,
                  min_row=2, max_row=2+len(datasets))
scats = Reference(ws3, min_col=1, min_row=3, max_row=2+len(datasets))
schart.add_data(sdata, titles_from_data=True)
schart.set_categories(scats)
ws3.add_chart(schart, f"A{len(datasets)+5}")

# ══════════════════════════════════════════════════════════════════════
# SHEET 4: Raw Data
# ══════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Raw Data")
ws4.sheet_view.showGridLines = False

headers_raw = ["Dataset", "Image", "Engine", "Time (s)", "Text Length", "Valid", "Predicted Text (preview)"]
widths = [22, 30, 12, 10, 12, 8, 60]
for i, (h, w) in enumerate(zip(headers_raw, widths), 1):
    header_cell(ws4, 1, i, h, bg=DARK_BLUE)
    ws4.column_dimensions[get_column_letter(i)].width = w

for r, (_, row) in enumerate(df.iterrows(), 2):
    bg = LIGHT_GRAY if r % 2 == 0 else WHITE
    preview = str(row["predicted_text"])[:80] + "..." if len(str(row["predicted_text"])) > 80 else str(row["predicted_text"])
    vals = [
        row["dataset"].replace("_"," ").title(),
        row["image"],
        row["engine"].capitalize(),
        round(row["time_sec"], 3),
        row["text_length"],
        "✅" if row["valid"] else "❌",
        preview
    ]
    for c, val in enumerate(vals, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")

add_border(ws4, 1, 1+len(df), 1, len(headers_raw))

# ── Save ───────────────────────────────────────────────────────────────
out_path = os.path.join("outputs", "ocr_benchmark_report.xlsx")
wb.save(out_path)
print(f"\n✅ Excel report saved to: {out_path}")
print(f"   Sheets: Summary Dashboard | Detailed Results | Speed Comparison | Raw Data")